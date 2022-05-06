from ast import Num
from tokenize import Number
from openpyxl import Workbook, load_workbook

from openpyxl.utils import get_column_letter

wb = load_workbook('./ExcelData.xlsx')

for sheet in wb:
    print(sheet)

ws = wb.active

def setHexadecimal(i):
    if i < 10:
        print(i)
    else:
        if i == 10:
            print("A")
        elif i == 11:
            print("B")
        elif i == 12:
            print("C")
        elif i == 13:
            print("D")
        elif i == 14:
            print("E")
        elif i == 15:
            print("F")

setHexadecimal(int(ws.cell(row = 2, column = 2).value))

